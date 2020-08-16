from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
import sys

wb = load_workbook("veterans.xlsx")
ws = wb['published']

start_index = column_index_from_string("D")
end_index = column_index_from_string("N")

idx = 10
while idx < 1000:
    if ws.cell(row=idx, column=column_index_from_string("C")).value is None:
        break
    idx += 1
published_len = idx-1

start_yr = int(ws.cell(row=published_len, column=column_index_from_string("C")).value)
end_yr = int(ws.cell(row=2, column=column_index_from_string("C")).value)

#total_authors_by_yr = {}
yr_and_authors = {}
cols = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

### parse and organize data ####
for i in range(2, published_len+1)[::-1]:
    for x in cols:
        if (ws.cell(row=i, column=column_index_from_string(x))).value:
            rawname = ws.cell(row=i, column=column_index_from_string(x)).value.strip()
            yr = int(ws.cell(row=i, column=column_index_from_string("C")).value)
            
            yr_and_authors[yr] = yr_and_authors.get(yr, [])

            if rawname not in yr_and_authors[yr]:
                #total_authors_by_yr[yr] = total_authors_by_yr.get(yr, 0) + 1
                yr_and_authors[yr].append(rawname)

### create 5 yr buckets ###
fiveyrbuckets = {}
for i in range(start_yr+5, end_yr+1):
    start = i - 5
    combined = []
    #print("\nfiveyrbuckets[%s] appended:" % i)
    while start < i:
        #print("\tadding yr %s" % start)
        for x in yr_and_authors[start]:
            combined.append(x)
        start += 1
    fiveyrbuckets[i] = combined

total_authors_by_yr = {new_list: [] for new_list in range(2015, 2020)}
new_women_by_yr = {new_list: [] for new_list in range(2015, 2020)}
new_men_by_yr = {new_list: [] for new_list in range(2015, 2020)}
veteran_women_by_yr = {new_list: [] for new_list in range(2015, 2020)}
veteran_men_by_yr = {new_list: [] for new_list in range(2015, 2020)}

for r in range(2, 901):
    yr = int(ws.cell(row=r, column=column_index_from_string("C")).value)
    if yr == start_yr + 4:
        break
    for c in cols:
        if ws.cell(row=r, column=column_index_from_string(c)).value:
            name = ws.cell(row=r, column=column_index_from_string(c)).value.strip()
            if name not in total_authors_by_yr[yr]:
                total_authors_by_yr[yr].append(name)
            if name in fiveyrbuckets[yr]:
                if name[0] == 'F':
                    if name not in veteran_women_by_yr[yr]:
                        veteran_women_by_yr[yr].append(name)
                elif name[0] == 'M':
                    if name not in veteran_men_by_yr[yr]:
                        veteran_men_by_yr[yr].append(name)
            else:
                if name[0] == 'F':
                    if name not in new_women_by_yr[yr]:
                        new_women_by_yr[yr].append(name)
                elif name[0] == 'M':
                    if name not in new_men_by_yr[yr]:
                        new_men_by_yr[yr].append(name)
                
for yr, tot in total_authors_by_yr.items():
    totnum = len(tot)
    print("%s:\n\ttotal authors: %s" % (yr, totnum))
    print("\tnew female authors: %s (%s%%)" % (len(new_women_by_yr[yr]), (len(new_women_by_yr[yr]) / totnum) * 100))
    print("\tnew male authors: %s (%s%%)" % (len(new_men_by_yr[yr]), (len(new_men_by_yr[yr]) / totnum) * 100))
    print("\tveteran female authors: %s (%s%%)" % (len(veteran_women_by_yr[yr]), (len(veteran_women_by_yr[yr]) / totnum) * 100 ))
    print("\tveteran male authors: %s (%s%%)" % (len(veteran_men_by_yr[yr]), (len(veteran_men_by_yr[yr]) / totnum) * 100))


