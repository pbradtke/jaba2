from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
import re
import json

wb = load_workbook("veterans.xlsx")
ws = wb['rejects']

total_authors_by_yr = {new_list: [] for new_list in range(2015, 2020)}
new_women_by_yr = {new_list: [] for new_list in range(2015, 2020)}
new_men_by_yr = {new_list: [] for new_list in range(2015, 2020)}
veteran_women_by_yr = {new_list: [] for new_list in range(2015, 2020)}
veteran_men_by_yr = {new_list: [] for new_list in range(2015, 2020)}
rejectcols = ['B', 'C', 'D', 'E', 'F', 'G', 'H']

for r in range(3, 966):
    yr = int(ws.cell(row=r, column=column_index_from_string("A")).value.split('-')[1])
    for c in rejectcols:
        if ws.cell(row=r, column=column_index_from_string(c)).value:
            name = ws.cell(row=r, column=column_index_from_string(c)).value.strip()
            if name not in total_authors_by_yr[yr]:
                total_authors_by_yr[yr].append(name)
            if name[-1] == 'N':
                if name[0] == 'F':
                    if name not in new_women_by_yr[yr]:
                        new_women_by_yr[yr].append(name)
                elif name[0] == 'M':
                    if name not in new_men_by_yr[yr]:
                        new_men_by_yr[yr].append(name)
            elif name[-1] == 'V':
                if name[0] == 'F':
                    if name not in veteran_women_by_yr[yr]:
                        veteran_women_by_yr[yr].append(name)
                elif name[0] == 'M':
                    if name not in veteran_men_by_yr[yr]:
                        veteran_men_by_yr[yr].append(name)

for yr, tot in total_authors_by_yr.items():
    totnum = len(tot)
    print("%s:\n\ttotal authors: %s" % (yr, totnum))
    print("\tnew female authors: %s (%s%%)" % (len(new_women_by_yr[yr]), (len(new_women_by_yr[yr]) / totnum) * 100))
    print("\tnew male authors: %s (%s%%)" % (len(new_men_by_yr[yr]), (len(new_men_by_yr[yr]) / totnum) * 100))
    print("\tveteran female authors: %s (%s%%)" % (len(veteran_women_by_yr[yr]), (len(veteran_women_by_yr[yr]) / totnum) * 100 ))
    print("\tveteran male authors: %s (%s%%)" % (len(veteran_men_by_yr[yr]), (len(veteran_men_by_yr[yr]) / totnum) * 100))






