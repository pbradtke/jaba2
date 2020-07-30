from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
import re
import json


wb = load_workbook("jabadata.xlsx")
ws = wb['published']

ws2 = wb['rejects']

start_index = column_index_from_string("D")
end_index = column_index_from_string("N")

### published ###
#published_len = len(ws["C"])
idx = 100
while idx < 1000:
    if ws.cell(row=idx, column=column_index_from_string("C")).value is None:
        break
    idx += 1
published_len = idx-1
start_yr = int(ws.cell(row=published_len, column=column_index_from_string("C")).value)
end_yr = int(ws.cell(row=2, column=column_index_from_string("C")).value)

### rejects ###
idx = 100
while idx < 1000:
    if ws2.cell(row=idx, column=column_index_from_string("A")).value is None:
        break
    idx += 1
#rejected_len = len(ws2["A"])
rejected_len = idx-1


target_authors_by_yr = {} #target authors in this case: rejected in year x and have not published in years x-5 through x-1
total_authors_by_yr = {}
yr_and_authors = {}
#first_authors_by_yr = {}
rejects_by_yr = {}

### parse and organize data ####
for i in range(2, published_len+1)[::-1]:
    for x in range(start_index, end_index+1):
        if (ws.cell(row=i, column=x)).value:
            rawname = ws.cell(row=i, column=x).value.strip()
            yr = int(ws.cell(row=i, column=column_index_from_string("C")).value)
            
            yr_and_authors[yr] = yr_and_authors.get(yr, [])

            if rawname not in yr_and_authors[yr]:
                total_authors_by_yr[yr] = total_authors_by_yr.get(yr, 0) + 1
                yr_and_authors[yr].append(rawname)

### create 5 yr buckets ###
fiveyrbuckets = {}
for i in range(start_yr+5, end_yr+1):
    start = i - 5
    combined = []
    #print("\nfiveyrbuckets[%s] appended:" % i)
    while start < i:
        #print("\tadding yr %s" % start)
        for x in yr_and_authors[start]:
            combined.append(x)
        start += 1
    fiveyrbuckets[i] = combined

### parse rejects ###
for i in range(3, rejected_len+1)[::-1]:
    for x in range(column_index_from_string("B"), column_index_from_string("B")+1):
        if (ws2.cell(row=i, column=x)).value:
            rawname = ws2.cell(row=i, column=x).value.strip()
            yr = int(ws2.cell(row=i, column=column_index_from_string("A")).value.split('-')[1])

            rejects_by_yr[yr] = rejects_by_yr.get(yr, [])

            if rawname not in rejects_by_yr[yr]:
                rejects_by_yr[yr].append(rawname)

### calculate target authors using five yr buckets ####
for i in range(start_yr+5, end_yr+1):
    for x in rejects_by_yr[i]:
        if x not in fiveyrbuckets[i]:
            target_authors_by_yr[i] = target_authors_by_yr.get(i,0) + 1

### print results ###
for k,v in total_authors_by_yr.items():
        if k < start_yr+5 or k not in target_authors_by_yr.keys():
            #print("%s: no matches found" % k)
            continue
        percent = float(target_authors_by_yr[k]/float(len(rejects_by_yr[k]))) * 100
        print("%s:\n \t%s rejected first authors not prublished in last 5 yrs\n \t%s total rejected first authors:\n \tpercent matching criteria: %s\n\n" % (k, target_authors_by_yr[k], len(rejects_by_yr[k]), percent))


